"""
data_utils.py
--------------
Fonctions de chargement, agrégation et normalisation du fichier
Palmares_Football_FM26_Complet_V3.xlsx pour le tableau de bord Streamlit.

Le classeur contient 60 feuilles :
    - 'Finalistes malheureux'   -> finalistes perdants en Ligue des Champions
    - 'Ligue des Champions'     -> vainqueurs de la C1 (avec colonne 'Pays')
    - 58 autres feuilles        -> un palmarès national par pays

Chaque feuille utilise le même code couleur de surlignage (fond de cellule)
défini manuellement dans Excel :
    - Rouge   (#FF0000) -> club aujourd'hui disparu
    - Vert    (#00B050) -> club mis en avant par l'utilisateur (surlignage vert)
    - Bleu    (#0070C0) -> club mis en avant par l'utilisateur (surlignage bleu)
    - Orange  (#FF9900) -> club mis en avant par l'utilisateur (surlignage orange)
    - (aucun)           -> pas de surlignage particulier

Ce module conserve la couleur exacte de chaque ligne afin de pouvoir
reproduire fidèlement la mise en forme d'origine dans l'interface.
"""

from __future__ import annotations

import openpyxl
import pandas as pd

SHEET_FINALISTES = "Finalistes malheureux"
SHEET_C1 = "Ligue des Champions"

# ------------------------------------------------------------------
# Légende des couleurs de surlignage rencontrées dans le classeur.
# Les codes sont ceux réellement présents dans le fichier (ARGB sans
# le préfixe alpha '00').
# ------------------------------------------------------------------
COLOR_LEGEND = {
    "FF0000": {"label": "Club disparu", "hex": "#FF0000"},
    "00B050": {"label": "Ancien champion dont le dernier titre est le plus ancien", "hex": "#00B050"},
    "0070C0": {"label": "Ancien champion évoluant au niveau de division le plus bas disponible", "hex": "#0070C0"},
    "FF9900": {"label": "Ancien champion au niveau de division le plus bas et dont le dernier titre est le plus ancien", "hex": "#FF9900"},
    "NONE": {"label": "Sans surlignage", "hex": "#FFFFFF"},
}

# En-tête (couleur du bandeau de titre) telle que définie dans Excel
HEADER_COLOR = "#1F4E78"

# ------------------------------------------------------------------
# Correspondance nom de pays (français, tel qu'utilisé dans les
# feuilles) -> code ISO-3166-1 alpha-3, nécessaire pour la carte
# choroplèthe. L'Écosse, le Pays de Galles et l'Irlande du Nord ne
# possèdent pas de code ISO propre : ils sont rattachés au polygone
# du Royaume-Uni (GBR) pour l'affichage cartographique uniquement,
# mais restent bien séparés dans les données et les tableaux.
# ------------------------------------------------------------------
COUNTRY_TO_ISO3 = {
    "Afrique du Sud": "ZAF", "Allemagne": "DEU", "Angleterre": "GBR",
    "Argentine": "ARG", "Australie": "AUS", "Autriche": "AUT",
    "Belgique": "BEL", "Biélorussie": "BLR", "Brésil": "BRA",
    "Bulgarie": "BGR", "Canada": "CAN", "Chili": "CHL", "Chine": "CHN",
    "Colombie": "COL", "Corée du Sud": "KOR", "Croatie": "HRV",
    "Danemark": "DNK", "Écosse": "GBR", "Égypte": "EGY",
    "Émirats Arabes Unis": "ARE", "Espagne": "ESP", "États-Unis": "USA",
    "Finlande": "FIN", "France": "FRA", "Gibraltar": "GIB",
    "Grèce": "GRC", "Hong Kong": "HKG", "Hongrie": "HUN", "Inde": "IND",
    "Indonésie": "IDN", "Irlande": "IRL", "Irlande du Nord": "GBR",
    "Islande": "ISL", "Israël": "ISR", "Italie": "ITA", "Japon": "JPN",
    "Lettonie": "LVA", "Lituanie": "LTU", "Malaisie": "MYS",
    "Mexique": "MEX", "Norvège": "NOR", "Pays de Galles": "GBR",
    "Pays-Bas": "NLD", "Pérou": "PER", "Pologne": "POL",
    "Portugal": "PRT", "République Tchèque": "CZE", "Roumanie": "ROU",
    "Russie": "RUS", "Serbie": "SRB", "Singapour": "SGP",
    "Slovaquie": "SVK", "Slovénie": "SVN", "Suède": "SWE",
    "Suisse": "CHE", "Turquie": "TUR", "Ukraine": "UKR",
    "Uruguay": "URY",
}


def _cell_color(cell) -> str:
    """Retourne le code couleur ARGB (sans alpha) d'une cellule, ou 'NONE'."""
    try:
        rgb = cell.fill.fgColor.rgb
    except AttributeError:
        rgb = None
    if not rgb or rgb == "00000000" or not isinstance(rgb, str):
        return "NONE"
    return rgb[-6:]  # on retire les 2 caractères d'alpha en tête


def load_workbook_dataframes(path: str) -> dict[str, pd.DataFrame]:
    """
    Lit toutes les feuilles du classeur et retourne un dict
    {nom_de_feuille: DataFrame}, chaque DataFrame incluant une colonne
    technique '_color' avec le code couleur de surlignage de la ligne.
    """
    wb = openpyxl.load_workbook(path, data_only=True)
    sheets = {}
    for sheetname in wb.sheetnames:
        ws = wb[sheetname]
        headers = [c.value for c in ws[1]]
        rows = []
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
            if row[0].value is None:
                continue
            values = [c.value for c in row]
            color = _cell_color(row[0])
            rows.append(values + [color])
        df = pd.DataFrame(rows, columns=headers + ["_color"])
        sheets[sheetname] = df
    return sheets


def build_unified_dataframe(sheets: dict[str, pd.DataFrame]) -> pd.DataFrame:
    """
    Construit un DataFrame unique à partir de toutes les feuilles,
    avec une colonne 'Source' (nom de la feuille d'origine) et une
    colonne 'Categorie' ('Ligue des Champions', 'Championnat national'
    ou 'Finalistes C1').

    Les colonnes sont normalisées :
        Club, Division_actuelle, Niveau_division, Annee_dernier_titre,
        Nb_titres, Nb_finales_perdues, Pays, Source, Categorie, _color
    """
    frames = []

    for sheetname, df in sheets.items():
        if df.empty:
            continue
        d = df.copy()

        if sheetname == SHEET_FINALISTES:
            d = d.rename(columns={
                "Nom de l'équipe": "Club",
                "Division actuelle": "Division_actuelle",
                "Niveau de la division": "Niveau_division",
                "Nombre de finales perdues": "Nb_finales_perdues",
            })
            d["Annee_dernier_titre"] = None
            d["Nb_titres"] = None
            d["Pays"] = None
            d["Categorie"] = "Finalistes C1"

        elif sheetname == SHEET_C1:
            d = d.rename(columns={
                "Nom de l'équipe": "Club",
                "Division actuelle": "Division_actuelle",
                "Niveau de la division": "Niveau_division",
                "Année du dernier titre": "Annee_dernier_titre",
                "Nombre de titres remportés": "Nb_titres",
                "Pays": "Pays",
            })
            d["Nb_finales_perdues"] = None
            d["Categorie"] = "Ligue des Champions"

        else:  # feuille pays
            d = d.rename(columns={
                "Nom de l'équipe": "Club",
                "Division actuelle": "Division_actuelle",
                "Niveau de la division": "Niveau_division",
                "Année du dernier titre dans l'élite": "Annee_dernier_titre",
                "Nombre de titres remportés": "Nb_titres",
            })
            d["Nb_finales_perdues"] = None
            d["Pays"] = sheetname
            d["Categorie"] = "Championnat national"

        d["Source"] = sheetname
        frames.append(d[[
            "Club", "Division_actuelle", "Niveau_division",
            "Annee_dernier_titre", "Nb_titres", "Nb_finales_perdues",
            "Pays", "Source", "Categorie", "_color",
        ]])

    unified = pd.concat(frames, ignore_index=True)
    unified["Iso3"] = unified["Pays"].map(COUNTRY_TO_ISO3)
    unified["Statut"] = unified["_color"].map(
        lambda c: COLOR_LEGEND.get(c, COLOR_LEGEND["NONE"])["label"]
    )
    return unified


def load_all(path: str):
    """Point d'entrée unique : renvoie (sheets_bruts, df_unifie)."""
    sheets = load_workbook_dataframes(path)
    unified = build_unified_dataframe(sheets)
    return sheets, unified
